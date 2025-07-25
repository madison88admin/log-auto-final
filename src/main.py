import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import io
import re
from typing import List
import json
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Canonical header mapping (adjust as needed to match frontend)
HEADER_MAPPING = {
    'CASE NOS': 'caseNos',
    'SA4 PO NO#': 'sa4PoNo',
    'CARTON PO NO#': 'cartonPoNo',
    'SAP STYLE NO': 'sapStyleNo',
    'STYLE NAME #': 'modelName',
    'S4 Material': 's4Material',
    'Material No#': 'materialNo',
    'COLOR': 'color',
    'Size': 'size',
    'Total QTY': 'totalQty',
    'CARTON': 'carton',
    'QTY / CARTON': 'unitsCrt',
    'TOTAL\nQTY': 'totalUnit',
    'N.W / ctn.': 'nwCtn',
    'TOTAL\nN.W.': 'totalNw',
    'G.W. / ctn': 'gwCtn',
    'TOTAL\nG.W.': 'totalGw',
    'MEAS.\nCM': 'measCm',
    'TOTAL\nCBM': 'totalCbm',
    'OS': 'OS', 'XS': 'XS', 'S': 'S', 'M': 'M', 'L': 'L', 'XL': 'XL', 'XXL': 'XXL',
    # Add any other variants as needed
}

# Choose the grouping key (adjust as needed)
GROUP_KEY = 'sa4PoNo'


# Normalize header
def normalize_header(header):
    return header.strip().replace('\n', ' ').replace('\r', '').replace('  ', ' ')

def extract_tables(file_bytes):
    xls = pd.ExcelFile(file_bytes)
    tables = []
    # Updated main_data_fields to match the main table headers
    main_data_fields = [
        'cartonNo', 'color', 's4Material', 'materialNo',
        'OS', 'XS', 'S', 'M', 'L', 'XL', 'XXL',
        'unitsCrt', 'totalUnit', 'totalNw', 'totalGw',
        'carton', 'Length', 'Width', 'Height',
        'cbm', 'totalCbm'
    ]
    for sheet_name in xls.sheet_names:
        df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
        header_indices = [i for i, row in df_raw.iterrows() if any(str(cell).strip() == "CASE NOS" for cell in row)]
        for idx, header_idx in enumerate(header_indices):
            start = header_idx
            end = header_indices[idx + 1] if idx + 1 < len(header_indices) else len(df_raw)
            table_rows = df_raw.iloc[start:end].values.tolist()
            if not table_rows or len(table_rows) < 2:
                continue
            header_row = [str(cell).strip() if cell is not None else "" for cell in table_rows[0]]
            mapped_keys = [HEADER_MAPPING.get(str(h).strip(), str(h).strip()) for h in header_row]
            print('Header row:', header_row)
            print('Mapped keys:', mapped_keys)
            # Find all contiguous non-blank rows after the header
            table_data = []
            size_names = ['OS', 'XS', 'S', 'M', 'L', 'XL', 'XXL']
            size_indices = {}
            for i, key in enumerate(mapped_keys):
                if key in size_names:
                    size_indices[key] = i
            for data_row in table_rows[1:]:
                # Stop at the first completely blank row
                if all((cell is None or str(cell).strip() == "" or str(cell).lower() == "nan") for cell in data_row):
                    break
                row_dict = {mapped_keys[i]: data_row[i] if i < len(data_row) else "" for i in range(len(header_row))}
                # Assign Length, Width, Height from R, S, T columns (Excel columns 18, 19, 20; Python indices 17, 18, 19)
                if len(data_row) > 17:
                    row_dict['Length'] = data_row[17]
                if len(data_row) > 18:
                    row_dict['Width'] = data_row[18]
                if len(data_row) > 19:
                    row_dict['Height'] = data_row[19]
                # Dynamically map each size column to its own key
                for size, idx in size_indices.items():
                    if idx is not None and idx < len(data_row):
                        row_dict[size] = data_row[idx]
                # Only include rows with at least one main data field filled
                if any(str(row_dict.get(f, '')).strip() not in ['', 'nan', 'None'] for f in main_data_fields):
                    table_data.append(row_dict)
            if table_data:
                print('Sample data row:', table_data[0])
            if not table_data:
                continue
            # Find the column index of 'CASE NOS' in the header row
            case_nos_col = None
            for col_idx, cell in enumerate(df_raw.iloc[header_idx]):
                if str(cell).strip() == "CASE NOS":
                    case_nos_col = col_idx
                    print(f"[DEBUG] Found 'CASE NOS' at column {case_nos_col}")
                    break
            # Extract model name from two rows above the CASE NOS header, same column
            model_name = ''
            if case_nos_col is not None and header_idx >= 2:
                model_name = str(df_raw.iloc[header_idx - 2, case_nos_col])
                print(f"[DEBUG] model_name (2 rows above header, col {case_nos_col}): {model_name}")
            sheet_name_val = next((row.get('sa4PoNo', '') or row.get('cartonNo', '') for row in table_data if row.get('sa4PoNo', '') or row.get('cartonNo', '')), f'Table{len(tables)+1}')
            safe_name = re.sub(r'[:\\/?*\[\]]', '_', str(sheet_name_val))[:31]
            tables.append({'rows': table_data, 'sheet_name': safe_name, 'model_name': model_name})
    return tables

def copy_worksheet(template_ws, target_ws):
    for row in template_ws.iter_rows():
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col, dim in template_ws.column_dimensions.items():
        target_ws.column_dimensions[col].width = dim.width
    for row, dim in template_ws.row_dimensions.items():
        target_ws.row_dimensions[row].height = dim.height
    for merged_range in template_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

def safe_float(val):
    try:
        f = float(val)
        if f != f:  # NaN check
            return 0
        return f
    except (ValueError, TypeError):
        return 0

def fill_template_with_data(ws, rows, group_name, model_name=None):
    if not rows:
        return
    # Fill header fields (adjust cell addresses as needed)
    first_row = rows[0]
    s4_hana_sku = str(first_row.get('s4Material', '') or first_row.get('E', ''))
    po_line_value = s4_hana_sku[:-2] if len(s4_hana_sku) > 2 else s4_hana_sku
    # Set PO-Line (D14) and Model # (D16) to trimmed S4 HANA SKU
    ws['D14'] = po_line_value
    ws['D16'] = po_line_value
    # Set SAP PO# (E14) and PO NO# to sheet name
    ws['E14'] = group_name
    # Always use the passed-in model_name for E16
    if model_name is not None:
        print(f"[DEBUG] Setting model_name in E16: {model_name}")
        ws['E16'] = model_name
    else:
        ws['E16'] = ''
    ws['E7'] = ''

    # Main table starts at row 20 (C20)
    main_table_start = 20
    num_data_rows = len(rows)
    template_data_rows = 1  # Assume template has 1 data row by default

    # --- Unmerge any merged cells in the main table data area (C20:V<end>) BEFORE inserting rows ---
    main_table_data_start = main_table_start
    main_table_data_end = main_table_start + num_data_rows + 20  # a safe buffer
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row >= main_table_data_start and
            merged_range.max_row <= main_table_data_end and
            merged_range.min_col >= 3 and
            merged_range.max_col <= 22
        ):
            ws.unmerge_cells(str(merged_range))

    if num_data_rows > template_data_rows:
        ws.insert_rows(main_table_start + 1, num_data_rows - template_data_rows)

    # --- MAIN TABLE: Propagate carton numbers and track ranges for merging (match App.tsx logic) ---
    effective_carton_nos = []
    last_carton_no = ''
    carton_row_ranges = {}
    for i, row in enumerate(rows):
        carton_no = str(row.get('caseNos', '')).strip()
        # Only propagate and track valid carton numbers (not empty, not 'nan')
        if carton_no and carton_no.lower() != 'nan':
            last_carton_no = carton_no
        effective_carton_nos.append(last_carton_no)
        row_num = main_table_start + i
        if last_carton_no and last_carton_no.lower() != 'nan':
            if last_carton_no not in carton_row_ranges:
                carton_row_ranges[last_carton_no] = {'start': row_num, 'end': row_num}
            else:
                carton_row_ranges[last_carton_no]['end'] = row_num

    # --- Write main table with split-carton OS logic and inline copy-down for D/E/F ---
    prev_color = ''
    prev_s4 = ''
    prev_ecc = ''
    for i, row in enumerate(rows):
        row_num = main_table_start + i
        carton_no = effective_carton_nos[i]
        # Only write Carton# for the first row in the group, else leave blank
        if carton_no and carton_no.lower() != 'nan' and carton_row_ranges.get(carton_no, {}).get('start') == row_num:
            ws.cell(row=row_num, column=3, value=carton_no)  # C: CASE NOS
        else:
            ws.cell(row=row_num, column=3, value='')

        # Color (D)
        color = str(row.get('color', '')).strip()
        if not color or color.lower() in ['nan', 'none']:
            color = prev_color
        ws.cell(row=row_num, column=4, value=color)
        prev_color = color

        # S4 HANA SKU (E)
        s4sku = str(row.get('s4Material', '')).strip()
        if not s4sku or s4sku.lower() in ['nan', 'none']:
            s4sku = prev_s4
        ws.cell(row=row_num, column=5, value=s4sku)
        prev_s4 = s4sku

        # ECC Material No (F)
        ecc = str(row.get('materialNo', '')).strip()
        if not ecc or ecc.lower() in ['nan', 'none']:
            ecc = prev_ecc
        ws.cell(row=row_num, column=6, value=ecc)
        prev_ecc = ecc

        # --- SPLIT CARTON LOGIC FOR OS COLUMN (split carton -> J, else L) ---
        if carton_no and carton_no.lower() != 'nan' and sum(1 for c in effective_carton_nos if c == carton_no) > 1:
            os_val = row.get('totalQty', '')  # Column J (index 9)
        else:
            os_val = row.get('unitsCrt', '')  # Column L (index 11)
        ws.cell(row=row_num, column=7, value=os_val)    # G: OS

        # Sizes H-M
        size_names = ['OS', 'XS', 'S', 'M', 'L', 'XL', 'XXL']
        for j, size in enumerate(size_names[1:]):  # skip OS, already filled
            ws.cell(row=row_num, column=8+j, value=row.get(size, ''))

        ws.cell(row=row_num, column=14, value=row.get('carton', ''))      # N: Carton
        ws.cell(row=row_num, column=15, value=row.get('totalUnit', ''))   # O: Total Unit
        # P: TOTAL N.W. (match source data exactly)
        ws.cell(row=row_num, column=16, value=row.get('totalNw', ''))
        # Q: TOTAL G.W. (match source data exactly)
        ws.cell(row=row_num, column=17, value=row.get('totalGw', ''))
        ws.cell(row=row_num, column=18, value=row.get('Length', ''))    # R: Length
        ws.cell(row=row_num, column=19, value=row.get('Width', ''))     # S: Width
        ws.cell(row=row_num, column=20, value=row.get('Height', ''))    # T: Height
        # U: CBM (rounded to nearest 10)
        cbm_value = safe_float(row.get('cbm', row.get('totalCbm', 0)))
        ws.cell(row=row_num, column=21, value=cbm_value)  # U: CBM
        ws.cell(row=row_num, column=22, value=cbm_value)  # V: TOTAL CBM

    # After writing the main table, round Columns P (16), Q (17), U (21), and V (22) to 3 decimal places for each data row
    for i in range(num_data_rows):
        row_num = main_table_start + i
        # P: Net Weight (rounded to 3 decimal places)
        net_weight = safe_float(ws.cell(row=row_num, column=16).value)
        ws.cell(row=row_num, column=16, value=round(net_weight, 3) if net_weight else 0)
        # Q: Gross Weight (rounded to 3 decimal places)
        gross_weight = safe_float(ws.cell(row=row_num, column=17).value)
        ws.cell(row=row_num, column=17, value=round(gross_weight, 3) if gross_weight else 0)
        # U: CBM (rounded to 3 decimal places)
        cbm_value = safe_float(ws.cell(row=row_num, column=21).value)
        ws.cell(row=row_num, column=21, value=round(cbm_value, 3) if cbm_value else 0)
        # V: TOTAL CBM (rounded to 3 decimal places)
        total_cbm_value = safe_float(ws.cell(row=row_num, column=22).value)
        ws.cell(row=row_num, column=22, value=round(total_cbm_value, 3) if total_cbm_value else 0)

    # Always write summary and color breakdown at fixed positions after the main table
    size_names = ['OS', 'XS', 'S', 'M', 'L', 'XL', 'XXL']
    main_table_start = 20
    num_data_rows = len(rows)
    summary_start_row = main_table_start + num_data_rows + 1
    summary_col = 4  # D
    value_col = 5    # E
    color_breakdown_col = 6  # F


    # --- Only merge Carton# and columns N–V (14–22) for split cartons
    for carton_no, rng in carton_row_ranges.items():
        if carton_no and carton_no.lower() != 'nan' and rng['end'] > rng['start']:
            ws.merge_cells(start_row=rng['start'], start_column=3, end_row=rng['end'], end_column=3)
            for col in range(14, 23):  # N–V (now includes N=14)
                ws.merge_cells(start_row=rng['start'], start_column=col, end_row=rng['end'], end_column=col)

    # --- SUMMARY AND COLOR BREAKDOWN (replicate frontend logic) ---
    # Always write summary and color breakdown at fixed positions after the main table
    size_names = ['OS', 'XS', 'S', 'M', 'L', 'XL', 'XXL']
    main_table_start = 20
    num_data_rows = len(rows)
    summary_start_row = main_table_start + num_data_rows + 1
    summary_col = 4  # D
    value_col = 5    # E
    color_breakdown_col = 6  # F

    # --- CLEAR TEMPLATE SUMMARY AND COLOR BREAKDOWN AREA ---

    # Determine where the main table ends
    main_table_end_row = main_table_start + num_data_rows - 1
    clear_start_row = main_table_end_row + 1
    clear_end_row = summary_start_row + 20
    for row in range(main_table_start, main_table_end_row + 1):
        for col in range(3, 23):  # C (3) to V (22)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    for row in range(clear_start_row, clear_end_row):
        for col in range(4, 15):  # D to N (or further if needed)
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.font = None
            cell.alignment = None
            cell.fill = PatternFill()
            cell.border = Border()

    # Unmerge any merged cells in this area
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row >= clear_start_row and
            merged_range.max_row <= clear_end_row and
            merged_range.min_col >= 4 and
            merged_range.max_col <= 15):
            ws.unmerge_cells(str(merged_range))

    # --- SUMMARY SECTION ---
    # Merge D and E for the summary name
    ws.merge_cells(start_row=summary_start_row, start_column=summary_col, end_row=summary_start_row, end_column=value_col)
    header_cell = ws.cell(row=summary_start_row, column=summary_col)
    header_cell.value = "Summary"
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.font = Font(bold=True)
    header_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Find the last Carton# value in the data rows
    carton_col = 3  # C
    last_carton_cell = ws.cell(row=main_table_start + num_data_rows - 1, column=carton_col).value
    processed_carton_value = last_carton_cell
    if last_carton_cell and isinstance(last_carton_cell, str) and '-' in last_carton_cell:
        parts = last_carton_cell.split('-')
        processed_carton_value = parts[-1].strip()

    # Sum columns P, Q, V, and N (Units/CRT) for the data rows
    total_net_weight = 0
    total_gross_weight = 0
    total_cbm = 0
    total_carton = 0  # Sum of Units/CRT (column N)
    last_units_crt = None
    for i in range(num_data_rows):
        row_num = main_table_start + i
        units_crt_val = ws.cell(row=row_num, column=14).value
        if units_crt_val not in [None, '', 'nan', 'None', 0, '0']:
            last_units_crt = units_crt_val
        total_carton += safe_float(last_units_crt)
    
    # --- FIX: Total Carton should be the count of unique, non-empty, non-'nan' carton numbers from the original data ---
    carton_numbers_set = set()
    for row in rows:
        carton_no = str(row.get('caseNos', '')).strip().lower()
        if carton_no and carton_no != 'nan':
            carton_numbers_set.add(carton_no)
    total_carton = len(carton_numbers_set)

    # --- SUMMARY TABLE CALCULATION AND WRITING ---
    main_table_end_row = main_table_start + len(rows) - 1
    # Check if there are any empty cells in column N (14)
    any_empty = any(ws.cell(row=row, column=14).value in [None, '', 'nan', 'None', 0, '0'] for row in range(main_table_start, main_table_end_row + 1))
    # Use copy-down sum if there are empty cells, else use direct sum
    if any_empty:
        total_carton, _ = copy_down_sum(ws, main_table_start, main_table_end_row, 14)
        print('DEBUG: Used copy-down sum for total_carton')
    else:
        total_carton = sum(
            int(ws.cell(row=row, column=14).value) if str(ws.cell(row=row, column=14).value).isdigit() else 0
            for row in range(main_table_start, main_table_end_row + 1)
        )
        print('DEBUG: Used direct sum for total_carton')
    # Sum of Net Weight (Column P = 16)
    total_net_weight = sum(
        safe_float(ws.cell(row=row, column=16).value)
        for row in range(main_table_start, main_table_end_row + 1)
    )
    # Sum of Gross Weight (Column Q = 17)
    total_gross_weight = sum(
        safe_float(ws.cell(row=row, column=17).value)
        for row in range(main_table_start, main_table_end_row + 1)
    )
    # Sum of Total CBM (Column V = 22)
    total_cbm = sum(
        safe_float(ws.cell(row=row, column=22).value)
        for row in range(main_table_start, main_table_end_row + 1)
    )

    # Write summary values (labels and values) below the summary header
    ws.cell(row=summary_start_row+1, column=summary_col, value='Total Carton')
    ws.cell(row=summary_start_row+1, column=value_col, value=int(total_carton))
    ws.cell(row=summary_start_row+2, column=summary_col, value='Total Net Weight')
    ws.cell(row=summary_start_row+2, column=value_col, value=round(total_net_weight, 3))
    ws.cell(row=summary_start_row+3, column=summary_col, value='Total Gross Weight')
    ws.cell(row=summary_start_row+3, column=value_col, value=round(total_gross_weight, 3))
    ws.cell(row=summary_start_row+4, column=summary_col, value='Total CBM')
    ws.cell(row=summary_start_row+4, column=value_col, value=round(total_cbm, 3))

    # --- COLOR BREAKDOWN SECTION (replicating frontend split-carton logic) ---
    # Place color breakdown headers to align with the summary header row
    color_breakdown_start_row = summary_start_row  # Align with summary header row
    color_breakdown_start_col = color_breakdown_col
    color_headers = ['Color'] + size_names + ['Total']
    border_style = Side(style='thin')
    header_border = Border(top=border_style, left=border_style, bottom=border_style, right=border_style)

    # Write headers with borders
    for i, header in enumerate(color_headers):
        cell = ws.cell(row=color_breakdown_start_row, column=color_breakdown_start_col + i, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    # --- NEW LOGIC: propagate carton numbers and build carton count map ---
    effective_carton_nos = []
    last_carton_no = ''
    for row in rows:
        carton_no = str(row.get('caseNos', '')).strip()
        if carton_no:
            last_carton_no = carton_no
        effective_carton_nos.append(last_carton_no)
    carton_count_map = {}
    for carton_no in effective_carton_nos:
        if not carton_no:
            continue
        carton_count_map[carton_no] = carton_count_map.get(carton_no, 0) + 1

    # --- NEW LOGIC: Color breakdown using worksheet values, propagating Units/CRT (N) down ---
    color_map = {}
    size_names = ['OS', 'XS', 'S', 'M', 'L', 'XL', 'XXL']
    size_col_indices = [7, 8, 9, 10, 11, 12, 13]  # G=7, H=8, ..., M=13 (1-based)
    color_col = 4  # D=4 (1-based)
    units_crt_col = 14  # N=14 (1-based)
    last_units_crt = None
    for i in range(num_data_rows):
        row_num = main_table_start + i
        color = ws.cell(row=row_num, column=color_col).value
        units_crt = ws.cell(row=row_num, column=units_crt_col).value
        if units_crt not in [None, '', 'nan', 'None', 0, '0']:
            last_units_crt = safe_float(units_crt)
        # If units_crt is empty, use the last non-empty value
        effective_units_crt = last_units_crt if last_units_crt is not None else 0
        if not color or str(color).strip().lower() in ['nan', 'none', '']:
            continue
        color = str(color).strip()
        if color not in color_map:
            color_map[color] = [0] * len(size_names)
        for j, size_col in enumerate(size_col_indices):
            size_val = safe_float(ws.cell(row=row_num, column=size_col).value)
            color_map[color][j] += size_val * effective_units_crt

    valid_colors = [color for color in color_map.keys() if color and color.lower() not in ['nan', 'none']]

    # Write color breakdown rows dynamically with borders
    for color_idx, color in enumerate(valid_colors):
        row_num = color_breakdown_start_row + 1 + color_idx
        color_cell = ws.cell(row=row_num, column=color_breakdown_start_col, value=color)
        color_cell.alignment = Alignment(horizontal='left')
        color_cell.border = header_border
        for i in range(len(size_names)):
            cell = ws.cell(row=row_num, column=color_breakdown_start_col + 1 + i, value=color_map[color][i])
            cell.alignment = Alignment(horizontal='center')
            cell.border = header_border
        total_value = sum(color_map[color])
        total_cell = ws.cell(row=row_num, column=color_breakdown_start_col + 8, value=total_value)
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='center')
        total_cell.border = header_border

    # Write the total row for color breakdown with borders
    total_row_num = color_breakdown_start_row + 1 + len(valid_colors)
    total_label_cell = ws.cell(row=total_row_num, column=color_breakdown_start_col, value="Total")
    total_label_cell.font = Font(bold=True)
    total_label_cell.alignment = Alignment(horizontal='left')
    total_label_cell.border = header_border
    for i in range(len(size_names)):
        size_total = sum(color_map[color][i] for color in valid_colors)
        cell = ws.cell(row=total_row_num, column=color_breakdown_start_col + 1 + i, value=size_total)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = header_border
    grand_total = sum(sum(color_map[color]) for color in valid_colors)
    grand_total_cell = ws.cell(row=total_row_num, column=color_breakdown_start_col + 8, value=grand_total)
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.alignment = Alignment(horizontal='center')
    grand_total_cell.border = header_border

    # --- Note: Removed dynamic thick border extension to avoid extra thick borders ---

    # --- Clear all borders in the top section (B2:W16) ---
    for row in range(2, 17):  # 2 to 16 inclusive
        for col in range(2, 24):  # B (2) to W (23)
            ws.cell(row=row, column=col).border = Border()

    # Clear borders in the area below the summary table, up to the start of the color breakdown or a safe buffer
    for row in range(summary_start_row + 5, summary_start_row + 15):  # Adjust the end as needed
        for col in range(2, 24):  # B (2) to W (23)
            ws.cell(row=row, column=col).border = Border()

    # --- Apply requested borders as per user instructions ---
    # 1. Outer Thin Border from D6 to D12
    thin_side = Side(style='thin')
    for row in range(6, 13):  # D6 to D12 (rows 6-12)
        for col in range(4, 5):  # D (4)
            border = Border()
            if row == 6:
                border = Border(top=thin_side, left=thin_side, right=thin_side)
            elif row == 12:
                border = Border(bottom=thin_side, left=thin_side, right=thin_side)
            else:
                border = Border(left=thin_side, right=thin_side)
            ws.cell(row=row, column=col).border = border

    # 2. Borders (Not Thick) in cells D13 to E16
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    for row in range(13, 17):  # 13 to 16
        for col in range(4, 6):  # D (4) to E (5)
            ws.cell(row=row, column=col).border = thin_border

    # 3. Thin Border in the merged cell F6 to G16
    for row in range(6, 17):
        for col in range(6, 8):  # F (6) to G (7)
            border = Border()
            if row == 6:
                border = Border(top=thin_side)
            if row == 16:
                border = Border(bottom=thin_side)
            if col == 6:
                border = Border(left=thin_side)
            if col == 7:
                border = Border(right=thin_side)
            # Combine borders for corners and edges
            if row == 6 and col == 6:
                border = Border(top=thin_side, left=thin_side)
            if row == 6 and col == 7:
                border = Border(top=thin_side, right=thin_side)
            if row == 16 and col == 6:
                border = Border(bottom=thin_side, left=thin_side)
            if row == 16 and col == 7:
                border = Border(bottom=thin_side, right=thin_side)
            if row > 6 and row < 16 and col == 6:
                border = Border(left=thin_side)
            if row > 6 and row < 16 and col == 7:
                border = Border(right=thin_side)
            ws.cell(row=row, column=col).border = border
    ws.merge_cells(start_row=6, start_column=6, end_row=16, end_column=7)

    # 4. Double line border from C3 to V3 (for merged cell containing "PACKING LIST")
    double_side = Side(style='double')
    
    # First, unmerge any existing merged cells in this area to avoid conflicts
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row == 3 and merged_range.max_row == 3 and 
            merged_range.min_col >= 3 and merged_range.max_col <= 22):
            ws.unmerge_cells(str(merged_range))
    
    # Now merge the cells C3:V3
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=22)
    
    # Apply complete double border to the merged cell
    # Top border for all cells in the merged range
    for col in range(3, 23):  # C (3) to V (22)
        cell = ws.cell(row=3, column=col)
        cell.border = Border(top=double_side)
    
    # Bottom border for all cells in the merged range
    for col in range(3, 23):  # C (3) to V (22)
        cell = ws.cell(row=3, column=col)
        current_border = cell.border
        cell.border = Border(
            top=current_border.top,
            left=current_border.left,
            right=current_border.right,
            bottom=double_side
        )
    
    # Left border for the first cell (C3)
    ws.cell(row=3, column=3).border = Border(
        top=double_side,
        left=double_side,
        bottom=double_side
    )
    
    # Right border for the last cell (V3)
    ws.cell(row=3, column=22).border = Border(
        top=double_side,
        right=double_side,
        bottom=double_side
    )

    # --- Apply thin borders to the main table (C20:V<end>) ---
    thin_side = Side(style='thin')
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    main_table_end_row = main_table_start + num_data_rows - 1
    for row in range(main_table_start, main_table_end_row + 1):
        for col in range(3, 23):  # C (3) to V (22)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(bold=False)  # Ensure not bold

    # --- Apply thin borders to the summary table (D/E, summary_start_row+1 to summary_start_row+4) ---
    for row in range(summary_start_row + 1, summary_start_row + 5):
        for col in range(4, 6):  # D (4) and E (5)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align

    # --- Apply thin borders to the color breakdown table (F to O, color_breakdown_start_row to total_row_num) ---
    for row in range(color_breakdown_start_row, total_row_num + 1):
        for col in range(6, 15):  # F (6) to O (14)
            ws.cell(row=row, column=col).border = thin_border

    # --- Draw a robust thick outer border around the entire report area ---
    top_row = 2
    bottom_row = total_row_num + 2  # last row of color breakdown
    left_col = 2  # Column B
    right_col = 23  # Column W

    # Top and bottom borders
    for col in range(left_col, right_col + 1):
        ws.cell(row=top_row, column=col).border = Border(
            top=Side(style='thick'),
            left=ws.cell(row=top_row, column=col).border.left,
            right=ws.cell(row=top_row, column=col).border.right,
            bottom=ws.cell(row=top_row, column=col).border.bottom
        )
        ws.cell(row=bottom_row, column=col).border = Border(
            bottom=Side(style='thick'),
            left=ws.cell(row=bottom_row, column=col).border.left,
            right=ws.cell(row=bottom_row, column=col).border.right,
            top=ws.cell(row=bottom_row, column=col).border.top
        )
        
    for row in range(bottom_row + 1, bottom_row + 20):  # Adjust 20 as a safe buffer
        for col in range(left_col, right_col + 1):
            ws.cell(row=row, column=col).border = Border()

# --- Clear borders above the thick top border (row 1 to top_row-1) ---
    for row in range(1, top_row):
        for col in range(left_col, right_col + 1):
            ws.cell(row=row, column=col).border = Border()

    # Left and right borders
    for row in range(top_row, bottom_row + 1):
        ws.cell(row=row, column=left_col).border = Border(
            left=Side(style='thick'),
            top=ws.cell(row=row, column=left_col).border.top,
            right=ws.cell(row=row, column=left_col).border.right,
            bottom=ws.cell(row=row, column=left_col).border.bottom
        )
        ws.cell(row=row, column=right_col).border = Border(
            right=Side(style='thick'),
            top=ws.cell(row=row, column=right_col).border.top,
            left=ws.cell(row=row, column=right_col).border.left,
            bottom=ws.cell(row=row, column=right_col).border.bottom
        )
        
    robust_copy_down(rows, ['color', 's4Material', 'materialNo'])

def robust_copy_down(rows, keys):
    last_values = {k: '' for k in keys}
    for row in rows:
        for k in keys:
            val = str(row.get(k, '')).strip()
            if not val or val.lower() in ['nan', 'none']:
                row[k] = last_values[k]
            else:
                last_values[k] = val

def copy_down_sum(ws, start_row, end_row, col):
    last_val = 0
    total = 0
    empty_found = False
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val not in [None, '', 'nan', 'None', 0, '0']:
            last_val = float(val)
        else:
            empty_found = True
        total += last_val
    return total, empty_found

# In generate_reports, skip NaN/Unknown group keys
@app.post("/generate-reports/")
async def generate_reports(
    file: UploadFile = File(...),
    table_keys: str = Form(None)  # JSON stringified list of keys
):
    print("LOG TEST: /generate-reports endpoint called")
    file_bytes = await file.read()
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    tables = extract_tables(io.BytesIO(file_bytes))
    template_wb = load_workbook("ReportTemplate.xlsx")
    template_ws = template_wb['Report']
    combined_wb = Workbook()
    if combined_wb.active is not None and combined_wb.active.title == "Sheet":
        combined_wb.remove(combined_wb.active)
    requested_keys = json.loads(table_keys) if table_keys else [t['sheet_name'] for t in tables]
    print("Available tables:", [t['sheet_name'] for t in tables])
    print("Requested keys:", requested_keys)
    normalized_tables = {str(t['sheet_name']).strip().lower(): t for t in tables}
    for key in requested_keys:
        if not key or str(key).lower() in ['nan', 'unknown']:
            continue
        norm_key = str(key).strip().lower()
        table = normalized_tables.get(norm_key)
        if not table or not table['rows']:
            continue
        ws = combined_wb.create_sheet(title=table['sheet_name'])
        copy_worksheet(template_ws, ws)
        fill_template_with_data(ws, table['rows'], table['sheet_name'], table['model_name'])
        # Copy C15–C20 from the input file to merged cell F6–G16
        try:
            input_sheet = None
            for sname in xls.sheet_names:
                if sname.strip().lower() == table['sheet_name'].strip().lower() or sname.strip().lower() in ['in', 'pk (2)']:
                    input_sheet = pd.read_excel(xls, sheet_name=sname, header=None, dtype=str)
                    break
            if input_sheet is not None:
                merged_text_lines = []
                for row_idx in range(14, 20):
                    val = input_sheet.iloc[row_idx, 2] if row_idx < len(input_sheet) else ''
                    if pd.notna(val) and str(val).strip():
                        merged_text_lines.append(str(val).strip())
                merged_text = '\n'.join(merged_text_lines)
                ws.merge_cells(start_row=6, start_column=6, end_row=16, end_column=7)
                ws.cell(row=6, column=6).value = merged_text
        except Exception as e:
            print(f"[DEBUG] Error copying C15-C20 to F6-G16: {e}")
    output = io.BytesIO()
    combined_wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=AllReports.xlsx"}
    )